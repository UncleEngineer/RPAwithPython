from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# excelfile = load_workbook()

import os

pythonfile_folder = os.getcwd() #current directory
excel_folder = os.path.join(pythonfile_folder,'excelfile')
excel_name = 'data-no-graph.xlsx'
excel_path = os.path.join(excel_folder, excel_name)

# print('Path:',excel_path)

excelfile = load_workbook(excel_path)
sheet = excelfile.active

print(sheet['A1'].value)

count_row = len(sheet['A'])
# print(sheet['A'])
print(count_row)
print('Max row:', sheet.max_row)
print('Max column:', sheet.max_column)

# ADD GRAPH TO CURRENT FILE
values = Reference(sheet,min_col=3,max_col=3, min_row=2,max_row=4)
chart = BarChart()
chart.title = 'ความสูงต้นไม้'
chart.y_axis.title = 'ความสูง'
chart.x_axis.title = 'ชนิดต้นไม้'
chart.legend = None

chart.add_data(values)
sheet.add_chart(chart,'E5')

cat = Reference(sheet, min_col=1,max_col=1, min_row=2, max_row=4)
chart.set_categories(cat)

excelfile.save(excel_path)